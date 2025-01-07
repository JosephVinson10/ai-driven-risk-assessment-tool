import pandas as pd
import random
import matplotlib.pyplot as plt
from fpdf import FPDF
import datetime
import openpyxl

# Step 1: Generate enhanced dummy data with additional risk factors
def generate_enhanced_dummy_data(num_entries=20):
    data = {
        "Risk ID": [f"RISK-{i+1}" for i in range(num_entries)],
        "Likelihood": [random.randint(1, 5) for _ in range(num_entries)],
        "Impact": [random.randint(1, 5) for _ in range(num_entries)],
        "Control Effectiveness": [random.randint(1, 5) for _ in range(num_entries)],
        "Detectability": [random.randint(1, 5) for _ in range(num_entries)],
        "Financial Impact ($)": [random.randint(1000, 10000) for _ in range(num_entries)],
        "Operational Impact": [random.randint(1, 5) for _ in range(num_entries)],
        "Reputational Impact": [random.randint(1, 5) for _ in range(num_entries)]
    }
    df = pd.DataFrame(data)
    return df

# Step 2: Create visualizations
def plot_risk_matrix(df):
    plt.figure(figsize=(8, 6))
    plt.scatter(df["Likelihood"], df["Impact"], c='red', s=100, alpha=0.6)
    plt.title("Risk Matrix (Likelihood vs. Impact)")
    plt.xlabel("Likelihood (1-5)")
    plt.ylabel("Impact (1-5)")
    plt.grid(True)
    plt.savefig("Risk_Matrix.png")
    plt.close()

def plot_risk_score_distribution(df):
    risk_scores = (df["Likelihood"] * df["Impact"]) / (df["Control Effectiveness"] + df["Detectability"])
    plt.figure(figsize=(8, 6))
    plt.hist(risk_scores, bins=5, color='blue', alpha=0.7, edgecolor='black')
    plt.title("Distribution of Risk Scores")
    plt.xlabel("Risk Score")
    plt.ylabel("Frequency")
    plt.grid(axis='y')
    plt.savefig("Risk_Score_Distribution.png")
    plt.close()

# Step 3: Export data to Excel
def export_to_excel(df, file_path="Enhanced_Risk_Assessment_Data.xlsx"):
    df.to_excel(file_path, index=False, engine='openpyxl')

# Step 4: Generate detailed, professional PDF report
def generate_professional_pdf_report(df, file_path="Professional_Risk_Assessment_Report.pdf"):
    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    
    # Cover Page
    pdf.add_page()
    pdf.set_font("Arial", style='B', size=16)
    pdf.cell(200, 10, txt="AI-Driven Risk Assessment Report", ln=True, align='C')
    pdf.ln(10)
    pdf.set_font("Arial", size=12)
    pdf.cell(200, 10, txt=f"Prepared by: Joseph Vinson", ln=True, align='C')
    pdf.cell(200, 10, txt=f"Date: {datetime.date.today()}", ln=True, align='C')
    pdf.ln(20)
    
    # Table of Contents
    pdf.set_font("Arial", style='B', size=14)
    pdf.cell(200, 10, txt="Table of Contents", ln=True)
    pdf.set_font("Arial", size=12)
    pdf.cell(200, 10, txt="1. Introduction", ln=True)
    pdf.cell(200, 10, txt="2. Methodology", ln=True)
    pdf.cell(200, 10, txt="3. Summary of Findings", ln=True)
    pdf.cell(200, 10, txt="4. Detailed Risk Assessment", ln=True)
    pdf.cell(200, 10, txt="5. Recommendations", ln=True)
    pdf.ln(10)
    
    # Introduction
    pdf.add_page()
    pdf.set_font("Arial", style='B', size=14)
    pdf.cell(200, 10, txt="1. Introduction", ln=True)
    pdf.set_font("Arial", size=12)
    pdf.multi_cell(0, 10, txt=(
        "This report presents the findings of an AI-driven risk assessment "
        "conducted on a set of dummy data representing various potential risks. "
        "The purpose of this report is to showcase the ability to identify, assess, "
        "and mitigate risks using automated tools and methodologies."
    ))
    pdf.ln(10)
    
    # Methodology
    pdf.set_font("Arial", style='B', size=14)
    pdf.cell(200, 10, txt="2. Methodology", ln=True)
    pdf.set_font("Arial", size=12)
    pdf.multi_cell(0, 10, txt=(
        "The risk assessment was conducted using a dataset generated with random values "
        "for the following factors:\n"
        "- Likelihood (1-5): The probability of the risk occurring.\n"
        "- Impact (1-5): The severity of the risk if it occurs.\n"
        "- Control Effectiveness (1-5): How well existing controls mitigate the risk.\n"
        "- Detectability (1-5): The ability to detect the risk before it causes harm.\n"
        "- Financial Impact ($): The potential financial loss.\n"
        "- Operational Impact (1-5): The effect on operations.\n"
        "- Reputational Impact (1-5): The potential damage to reputation."
    ))
    pdf.ln(10)
    
    # Summary of Findings
    pdf.set_font("Arial", style='B', size=14)
    pdf.cell(200, 10, txt="3. Summary of Findings", ln=True)
    pdf.set_font("Arial", size=12)
    high_risks = df[df["Likelihood"] * df["Impact"] > 16].shape[0]
    medium_risks = df[(df["Likelihood"] * df["Impact"] <= 16) & (df["Likelihood"] * df["Impact"] > 8)].shape[0]
    low_risks = df[df["Likelihood"] * df["Impact"] <= 8].shape[0]
    pdf.cell(200, 10, txt=f"Total Risks Assessed: {df.shape[0]}", ln=True)
    pdf.cell(200, 10, txt=f"High Risks: {high_risks}", ln=True)
    pdf.cell(200, 10, txt=f"Medium Risks: {medium_risks}", ln=True)
    pdf.cell(200, 10, txt=f"Low Risks: {low_risks}", ln=True)
    pdf.ln(10)
    
    # Detailed Risk Assessment
    pdf.set_font("Arial", style='B', size=14)
    pdf.cell(200, 10, txt="4. Detailed Risk Assessment", ln=True)
    pdf.set_font("Arial", size=12)
    
    for i, row in df.iterrows():
        pdf.cell(200, 10, txt=f"Risk ID: {row['Risk ID']}", ln=True)
        pdf.cell(200, 10, txt=f"Likelihood: {row['Likelihood']}, Impact: {row['Impact']}", ln=True)
        pdf.cell(200, 10, txt=f"Control Effectiveness: {row['Control Effectiveness']}, Detectability: {row['Detectability']}", ln=True)
        pdf.cell(200, 10, txt=f"Financial Impact: ${row['Financial Impact ($)']}, Operational Impact: {row['Operational Impact']}, Reputational Impact: {row['Reputational Impact']}", ln=True)
        pdf.ln(5)
    
    # Recommendations
    pdf.set_font("Arial", style='B', size=14)
    pdf.cell(200, 10, txt="5. Recommendations", ln=True)
    pdf.set_font("Arial", size=12)
    pdf.multi_cell(0, 10, txt=(
        "Based on the findings, it is recommended to:\n"
        "- Focus on mitigating high-risk items by improving control measures and detectability.\n"
        "- Conduct regular reviews of operational and financial impacts.\n"
        "- Enhance reputational risk management strategies."
    ))
    
    pdf.output(file_path)

# Main function to run the process
if __name__ == "__main__":
    # Generate dummy data
    enhanced_data = generate_enhanced_dummy_data(20)
    
    # Create visualizations
    plot_risk_matrix(enhanced_data)
    plot_risk_score_distribution(enhanced_data)
    
    # Export data to Excel
    export_to_excel(enhanced_data)
    
    # Generate professional PDF report
    generate_professional_pdf_report(enhanced_data)
